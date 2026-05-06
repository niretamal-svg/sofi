import os
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

def extract_docx(file_path, out_file):
    out_file.write(f"--- Extracting {file_path} ---\n")
    doc = Document(file_path)
    for para in doc.paragraphs:
        if para.text.strip():
            out_file.write(para.text + "\n")
    out_file.write("\n\n")

def extract_xlsx(file_path, out_file):
    out_file.write(f"--- Extracting {file_path} ---\n")
    wb = load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        out_file.write(f"Sheet: {sheet_name}\n")
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
            if row_str.replace(" | ", "").strip():
                out_file.write(row_str + "\n")
    out_file.write("\n\n")

def extract_pdf(file_path, out_file):
    out_file.write(f"--- Extracting {file_path} ---\n")
    reader = PdfReader(file_path)
    for i, page in enumerate(reader.pages):
        out_file.write(f"Page {i+1}:\n")
        text = page.extract_text()
        if text:
            out_file.write(text + "\n")
    out_file.write("\n\n")

if __name__ == "__main__":
    base_dir = r"c:\Users\nikolas\Desktop\sofi\sofi-app-20260501T183203Z-3-001\sofi-app\docs-extracted\Docs. Reunión Jueves 30-04"
    
    with open("extracted_data.txt", "w", encoding="utf-8") as f:
        extract_docx(os.path.join(base_dir, "SOFI_Informe_Tecnico_API.docx"), f)
        extract_xlsx(os.path.join(base_dir, "Modelo_Datos_Sofi_Publ.xlsx"), f)
        extract_xlsx(os.path.join(base_dir, "Modelo_NoSQL_Sofi_Publ.xlsx"), f)
        extract_pdf(os.path.join(base_dir, "Plan de Trabajo.pdf"), f)
